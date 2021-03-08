import sys
from openpyxl import load_workbook


import user_data as ud
from xl_helpers import check_file


def xlsx_reader(header_row, data_row, data_col, fn, verbose=False):
    workbook = load_workbook(filename=fn, data_only=True)
    sheet = workbook.active
    header = list(list(sheet.iter_rows(
        min_row=header_row, max_row=header_row, min_col=2))[0])

    result = []
    buff_str = ''
    line_count = 0
    stop = False

    for row in sheet.iter_rows(min_row=data_row, min_col=data_col):
        if stop:
            break

        line_count += 1
        buffer = {}
        for current_cell in row:
            e = header[current_cell.column-data_col].value
            buffer[e] = current_cell.value
            if buffer[e] == None:
                buffer[e] = '-'
                if (e == ud.date_stamp):
                    stop = True
                    break
            buff_str += '{0:<25s}'.format(str(buffer[e])[:24])
        buff_str += '\n'
        result.append(buffer)
    if verbose:
        print('\n\n' + buff_str)
        print('\nProcessed {} lines\n\n'.format(line_count))

    return result


def main_xl():
    current_month = '{:02d}'.format(int(input("Please enter month:")))
    xl_name = input("Please enter xl file name:")
    file_name = ud.cwd + ud.fd + ud.data_path + \
        ud.fd + current_month + ud.fd + xl_name
    result = xlsx_reader(ud.xl_header_row,
                         ud.xl_data_row,
                         ud.xl_data_row - ud.xl_header_row,
                         check_file(file_name))
    return result


if __name__ == "__main__":
    if len(sys.argv) == 1:
        result = main_xl()
        print('List of lenght: ' + str(len(result)))
    elif sys.argv[1] == "hello":
        print("hello world!")
    else:
        print(sys.argv)
