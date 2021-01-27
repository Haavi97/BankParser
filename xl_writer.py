from openpyxl import load_workbook


import user_data as ud


def xl_writer(current_month, data,  fn=ud.fn, fmap='csv'):
    current_month = str(current_month)
    workbook = load_workbook(filename=fn)
    try:
        sheet = workbook[current_month]
    except KeyError:
        print("Creating worksheet: \"" + current_month + "\"")
        workbook.create_sheet(current_month)
        sheet = workbook[current_month]
        sheet.append(ud.xl_month_fields)
    for element in data:
        f_append(sheet, element, fmap)
    workbook.save(fn)


def f_append(sheet, element, fmap):
    row = []
    if fmap == 'csv':
        fmaps = ud.xl_csv_map
    elif fmap == 'xl':
        fmaps = ud.xl_xl_map
    elif fmap == 'csv_def':
        fmaps = ud.xl_csv_def_map
    elif fmap == 'xl_def':
        fmaps = ud.xl_xl_def_map
    else:
        print("Wrong fmap")
    for e in ud.xl_month_fields:
        try:
            current = element[fmaps[e]]
        except:
            current = ""
        row.append(current)
    sheet.append(row)
